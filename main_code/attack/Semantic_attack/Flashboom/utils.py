import ast
import csv
import os
import re
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font


def get_txt_content_as_str(file_dir)->str:
    with open(file_dir, 'r', encoding='utf-8') as f:
        content = f.read()

    return content

def get_txt_content_as_lines(file_dir)->list:
    with open(file_dir, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    return lines


def get_explain_by_id(id:int)->str:
    explain_dir = 'data/vuln-10/explanation'
    for file in os.listdir(explain_dir):
        if file.startswith(str(id)):
            return get_txt_content_as_str(os.path.join(explain_dir, file))

    raise IndexError(f'file start with {id} not found')

def get_code_and_filename_list(code_dir):
    code_list = []

    for root, dirs, files in os.walk(code_dir):
        for file in files:
            if file.endswith(".sol"):
                file_full_path = os.path.join(root, file)
                with open(file_full_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                    code_list.append((content, file))
    
    return code_list


def get_code_list(code_dir):
    code_list = []

    for root, dirs, files in os.walk(code_dir):
        for file in files:
            if file.endswith(".sol"):
                file_full_path = os.path.join(root, file)
                with open(file_full_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                    code_list.append(content)
    
    return code_list


'''
return indexs of lines which start with keys
'''
def indexs_of_lines_start_with(lines, keys: tuple)->list:
    indexs = []
    for i,line in enumerate(lines, 0):
        if line.strip().startswith(keys):
            indexs.append(i)
    return indexs


def write_csv(header, data, csv_filename):
    with open(csv_filename, 'w', newline='') as csv_file:
        writer = csv.writer(csv_file)

        # 写入列名
        writer.writerow(header)

        # 写入数据行
        for key, values in data.items():
            writer.writerow([key] + values)

    print(f"CSV file {csv_filename} has been created successfully.")

def write_xlsx(header, data, xlsx_filename):
    # 创建一个新的 Excel 工作簿和工作表
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(header)

    for key, values in data.items():
        ws.append([key] + values)

    # 设置字体，垂直居中+左对齐，自动换行
    font = Font(name='Microsoft YaHei Mono')
    alignment = Alignment(vertical='center', horizontal='left', wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.font = font
            cell.alignment = alignment

    # 设置第3列开始的列宽为30
    for col in range(3, ws.max_column + 1):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 30

    # 保存文件
    wb.save(xlsx_filename)
    print(f"xlsx file {xlsx_filename} has been created successfully.")


def get_function_ranges_of_source_code(source_path, source_format):
    if source_format == '.cpp':
        return get_function_ranges_of_cpp(source_path)
    elif source_format == '.sol':
        return get_function_ranges_of_solidity(source_path)
    elif source_format == '.py':
        return get_function_ranges_of_python(source_path)
    else:
        raise ValueError('unknown source file format: ', source_format)
'''
return a dict, {contract_name.function_name:[start_line, end_line]}
'''
def get_function_ranges_of_solidity(sol_source_path)->dict:
    with open(sol_source_path, 'r') as file:
        lines = file.readlines()

    contract_function_ranges = {}
    current_contract = None
    current_function = None
    function_start_line = None

    contract_pattern = re.compile(r'^\s*(contract|interface|library)\s+(\w+)')
    function_pattern = re.compile(r'^\s*function\s+(\w+)?\s*\(')
    # TODO: modifier? event?
    brace_stack = []
    once_left_brace = False

    for i, line in enumerate(lines, 0):
        # Check if the line defines a new contract, interface, or library
        contract_match = contract_pattern.match(line)
        if contract_match:
            current_contract = contract_match.group(2)
            continue

        # Check if the line defines a new function
        function_match = function_pattern.match(line)
        if function_match:
            if current_function and current_contract:
                # If there's an ongoing function, store its range
                contract_function_ranges[f"{current_contract}.{current_function}"] = [function_start_line, i - 1]
                brace_stack = []
                once_left_brace = False
            current_function = function_match.group(1) if function_match.group(1) else 'fallback'
            function_start_line = i
            
        # Check if the function ends (by checking for a closing brace '}') 
        if current_function:
            for ch in line:
                if ch == '{':
                    brace_stack.append('{')
                    once_left_brace = True
                elif ch == '}':
                    if len(brace_stack) > 0 and brace_stack[-1] == '{':
                        brace_stack.pop(-1)
                    else:
                        print(f'unclosed brace or just definition at line {i}, {sol_source_path}')
                        # drop function definitions
                        current_function = None
                        once_left_brace = False
                        break
            if len(brace_stack) == 0 and once_left_brace:
                contract_function_ranges[f"{current_contract}.{current_function}"] = [function_start_line, i]
                current_function = None
                function_start_line = None
                brace_stack = []
                once_left_brace = False

    return contract_function_ranges



def get_function_ranges_of_cpp(file_path):
    with open(file_path, 'r') as f:
        lines = f.readlines()
    

    # 定义正则表达式匹配函数定义，排除控制结构（if, for, while 等）

    class_pattern = re.compile(r'^\s*class\s+(\w+)\s*\{')
    #function_pattern = re.compile(r'^\s*((\w|<|>)+)\s*(\w+)\s*\(.*\)\s*\{')  # 匹配函数名
    function_pattern = re.compile(r'^\s*([\w\s:<>,*&]+)\s+(\w+)\s*\(.*\)\s*\{')
    control_structures = re.compile(r'^\s*(if|for|while|switch|else)\s*\(.*\)\s*\{')
    else_if_pattern = re.compile(r'\s*else\s*if\s*\(.+\).*')
    comment_pattern = re.compile(r'^\s*(//|/\*|\*)')

    class_function_ranges = {}
    current_class = None
    current_function = None
    function_start_line = None

    brace_stack = []
    once_left_brace = False

    for i, line in enumerate(lines, 0):
        class_match = class_pattern.match(line)
        if class_match:
            current_class = class_match.group(1)
            continue
        # 检查是否匹配函数定义
        function_match = function_pattern.match(line)
        if function_match and not (else_if_pattern.match(line) or control_structures.match(line) or comment_pattern.match(line)):
            # 函数内部又出现函数定义，说明是
            # 1. 前一个函数只有定义，没有实现
            # 2. 前一个函数内部还有函数
            # 直接忽略第二个函数，当做普通行处理
            if current_function:
                #class_function_ranges[f'{current_class}.{current_function}'] = [function_start_line, i-1]
                #current_function = None
                print('-----')
                #print(f'{current_class}.{current_function}: {class_function_ranges[f"{current_class}.{current_function}"]}')
                print('group0:', function_match.group(0))
                print('group1:', function_match.group(1))
                print('group2:', function_match.group(2))
                #raise ValueError('stop early!!!')
                print(('stop early!!!unknown pattern maybe'))
                print('-------')
                pass
            # print('match line: ', line)
            else:
                current_function = function_match.group(2)  # 函数名
                function_start_line = i
                brace_stack = []
                once_left_brace = False
            
            
        # Check if the function ends (by checking for a closing brace '}') 
        if current_function:
            #print(i, line, brace_stack)
            for ch in line:
                if ch == '{':
                    brace_stack.append('{')
                    once_left_brace = True
                elif ch == '}':
                    if len(brace_stack) > 0 and brace_stack[-1] == '{':
                        brace_stack.pop(-1)
                    else:
                        print(f'unclosed brace or just definition at line {i}, {file_path}')
                        # drop function definitions
                        current_function = None
                        once_left_brace = False
                        break
            if len(brace_stack) == 0 and once_left_brace:
                class_function_ranges[f"{current_class}.{current_function}"] = [function_start_line, i]
                current_function = None
                function_start_line = None
                brace_stack = []
                once_left_brace = False
                #print(i, line)

    return class_function_ranges


def get_function_ranges_of_python(source_path):
    # 结果字典
    function_ranges = {}

    # 读取源码文件
    with open(source_path, "r", encoding="utf-8") as file:
        source_code = file.read()

    # 解析源码
    tree = ast.parse(source_code)

    # 遍历 AST 树
    for node in ast.walk(tree):
        if isinstance(node, ast.FunctionDef) or isinstance(node, ast.AsyncFunctionDef):
            # 检查函数的父节点是否是类
            parent_class = None
            for parent in ast.walk(tree):
                if isinstance(parent, ast.ClassDef):
                    if node in parent.body:
                        parent_class = parent.name
                        break
            
            # 确定类名
            class_name = parent_class if parent_class else "noclass"

            # 计算行号范围
            start_line = node.lineno-1
            end_line = node.end_lineno if hasattr(node, "end_lineno") else start_line

            # 构造 key
            func_name = f"{class_name}.{node.name}"
            function_ranges[func_name] = [start_line, end_line]

    return function_ranges

'''
return a list, linewise attention
'''
def linewise_attention_from_csv(csv_path)->list:
    if not os.path.exists(csv_path):
        print('csv not found: ', csv_path)
        return []
    matrix = np.loadtxt(csv_path, delimiter=',', dtype=float)
    #print('matrix:', matrix)

    if matrix.ndim == 2:
        sum_of_lines = matrix.sum(axis=1)
    elif matrix.ndim == 1:
        sum_of_lines = [matrix.sum(axis=0)]
    return sum_of_lines.tolist()

'''
save the contents of elements in functions_need from sol_source_path to output_path_format
output_path: sol_source_path_name-{function_name}.sol
'''
def save_function_content(functions_need:list, source_path, source_format, output_path_format):
    with open(source_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()
    function_ranges = get_function_ranges_of_source_code(source_path, source_format)
    for function_name in functions_need:
        se = function_ranges[function_name]
        start_line = se[0]
        end_line = se[1]
        lines_need = lines[start_line:end_line+1]
        output_path = output_path_format.format(function_name=function_name)
        with open(output_path, 'w', encoding='utf-8') as of:
            of.writelines(lines_need)


import os
import warnings
from contextlib import redirect_stdout

def suppress_print_and_warnings(func, *args, **kwargs):
    # 保存原来的警告过滤器
    original_filters = warnings.filters[:]
    
    # 将 stdout 重定向到 os.devnull（即 "空设备"）
    with open(os.devnull, 'w') as fnull:
        with redirect_stdout(fnull):
            # 临时禁用所有警告
            warnings.filterwarnings("ignore")
            result = func(*args, **kwargs)
    
    # 恢复原来的警告过滤器
    warnings.filters = original_filters
    return result


def get_ext_by_dataset(dataset):
    ext = None
    match dataset:
        case 'smartbugs-collection': ext = '.sol'
        case 'big-vul-100': ext = '.cpp'
        case 'cvefixes-100': ext = '.py'

        case 'messiq_dataset': ext = '.sol'
        case 'leetcode_cpp': ext = '.cpp'
        case 'leetcode_python': ext = '.py'
    if not ext:
        raise ValueError('unknown dataset', dataset)
    return ext


def get_vuln_func_name_and_range(vuln_dataset, case_name, ext):
    source_before_path = f'data/{vuln_dataset}/code/{case_name}{ext}'
    func_range_before = get_function_ranges_of_source_code(source_path=source_before_path, source_format=ext)
    if ext == '.sol':
        # find vuln func name in explanation
        explan_path = f'data/{vuln_dataset}/explanation/{case_name}.txt'
        # explan_str = get_txt_content_as_str(explan_path)
        # line_no_pattern = r'^The vulnerability lies in line (\d+):'
        # line_nos = re.findall(line_no_pattern, explan_str)
        # line_nos = [int(n) for n in line_nos]
        
        # vuln_func_names = set()
        # for line_no in line_nos:
        #     for func in func_range_before:
        #         func_start_line_no = func_range_before[func][0]
        #         func_end_line_no = func_range_before[func][1]
        #         if line_no <= func_end_line_no and line_no >= func_start_line_no:
        #             vuln_func_names.add(func)
        #             break
        # 提取explan中的所在行内容，去source中找它在哪一行
        vuln_func_names = set()
        code_lines_before = get_txt_content_as_lines(source_before_path)
        explan_lines = get_txt_content_as_lines(explan_path)
        for explan_line in explan_lines:
            if 'The vulnerability lies in line' in explan_line: 
                content = explan_line.split(':')[-1].strip()
                content = content.split('//')[0]
                content = content.split('/*')[0]
                #print(content)
                found = False
                for func in func_range_before:
                    func_start_line_no = func_range_before[func][0]
                    func_end_line_no = func_range_before[func][1]
                    func_lines = code_lines_before[func_start_line_no:func_end_line_no+1]
                    for func_line in func_lines:
                        if content in func_line:
                            vuln_func_names.add(func)
                            found = True
                            break
                    if found: break

    elif ext == '.cpp' or ext == '.py':
        # only one func
        
        vuln_func_names = set(func_range_before.keys())

    vuln_func_names = list(vuln_func_names)

    return vuln_func_names[0], func_range_before[vuln_func_names[0]]